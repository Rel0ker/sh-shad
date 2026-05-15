"""
Заполнение шаблона приказа об оплате часов замещения (как в «Замены март …xlsx»).
Шапка и оформление листа сохраняются из шаблона; заполняется таблица и строка «Всего часов».
"""

from __future__ import annotations

import io
import sys
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_MONTH_GENITIVE = (
    "",
    "январе",
    "феврале",
    "марте",
    "апреле",
    "мае",
    "июне",
    "июле",
    "августе",
    "сентябре",
    "октябре",
    "ноябре",
    "декабре",
)


def _template_path() -> Path:
    if getattr(sys, "frozen", False):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).resolve().parent.parent.parent
    return base / "data" / "templates" / "zameny_zp_template.xlsx"


def _find_total_row(ws) -> int:
    for r in range(15, min(ws.max_row + 1, 400)):
        v = ws.cell(r, 3).value
        if v is not None and str(v).strip() == "Всего часов":
            return r
    raise ValueError("В шаблоне не найдена строка «Всего часов» (колонка C).")


def _unmerge_data_block(ws, first_row: int, last_row: int) -> None:
    to_remove: list[str] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= first_row and mr.max_row <= last_row:
            to_remove.append(str(mr))
    for s in to_remove:
        ws.unmerge_cells(s)


def _merge_gi(ws, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 9) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)


def _normalize_day_date(raw: Any) -> str | None:
    """В SQLite дата может прийти строкой, date или datetime — нужен YYYY-MM-DD."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    if not s:
        return None
    if " " in s:
        s = s.split()[0]
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def aggregate_zp_rows(changes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Одна строка приказа на пару «заменяющий + отсутствующий» (как в типовой ведомости ЗП).

    Колонка «Ф. И. О, учителя» (C) — учитель, **получающий** оплату за замену (заменяющий).
    «Ф. И. О. замещаемого» (F) — отсутствующий.

    Часы = число сохранённых уроков (строк в базе) за период по этой паре.
    Раньше строки дробились по тексту примечания — в приказе это давало лишние строки и
    «рвало» часы; теперь причины по одной паре объединяются.
    """
    groups: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {"hours": 0, "dates": set(), "reason_cnt": Counter()}
    )
    for c in changes:
        repl = (c.get("replacement_fio") or "").strip()
        absent = (c.get("absent_fio") or "").strip()
        if not repl or not absent:
            continue
        day = _normalize_day_date(c.get("day_date"))
        if not day:
            continue
        note = (c.get("note") or "").strip()
        reason_token = note if note else "Приказ"
        key = (repl, absent)
        g = groups[key]
        g["hours"] += 1
        g["dates"].add(day)
        g["reason_cnt"][reason_token] += 1

    rows: list[dict[str, Any]] = []
    for (repl, absent), g in sorted(
        groups.items(),
        key=lambda kv: (kv[0][0].lower(), kv[0][1].lower()),
    ):
        dates_sorted = sorted(g["dates"])
        parts = []
        for d in dates_sorted:
            dt = datetime.strptime(d, "%Y-%m-%d").date()
            parts.append(dt.strftime("%d.%m."))
        dates_str = ", ".join(parts)
        cnt: Counter[str] = g["reason_cnt"]
        mx = max(cnt.values()) if cnt else 0
        cands = [k for k, v in cnt.items() if v == mx] if mx else ["Приказ"]
        reason = cands[0] if len(cands) == 1 else "; ".join(sorted(cands))
        rows.append(
            {
                "replacement": repl,
                "absent": absent,
                "hours": int(g["hours"]),
                "dates": dates_str,
                "reason": reason,
            }
        )
    return rows


def _build_title(date_from: str, date_to: str) -> str:
    d1 = datetime.strptime(date_from, "%Y-%m-%d").date()
    d2 = datetime.strptime(date_to, "%Y-%m-%d").date()
    if d1.month == d2.month and d1.year == d2.year:
        m = _MONTH_GENITIVE[d1.month]
        return (
            f" Об оплате часов, выполненных в порядке замещения отсутствующих учителей "
            f"в {m} ({d1.strftime('%d.%m')}– {d2.strftime('%d.%m')}) {d2.year} г."
        )
    return (
        f" Об оплате часов, выполненных в порядке замещения отсутствующих учителей "
        f"за период с {d1.strftime('%d.%m.%Y')} по {d2.strftime('%d.%m.%Y')} г."
    )


def build_prikaz_xlsx(
    changes: list[dict[str, Any]],
    date_from: str,
    date_to: str,
    *,
    order_date: date | None = None,
) -> bytes:
    tpl = _template_path()
    if not tpl.is_file():
        raise FileNotFoundError(
            f"Не найден шаблон {tpl}. Положите файл zameny_zp_template.xlsx в data/templates/."
        )

    order_date = order_date or date.today()
    groups = aggregate_zp_rows(changes)

    buf = io.BytesIO(tpl.read_bytes())
    wb = load_workbook(buf)
    ws = wb.active

    total_row = _find_total_row(ws)
    data_start = 17
    old_capacity = total_row - data_start
    n = len(groups)

    if n > old_capacity:
        ws.insert_rows(total_row, amount=n - old_capacity)
        total_row += n - old_capacity

    last_data_row = total_row - 1
    _unmerge_data_block(ws, data_start, last_data_row)

    for r in range(data_start, total_row):
        for c in range(2, 10):
            ws.cell(r, c).value = None

    style_src = data_start
    for i, row in enumerate(groups, start=1):
        r = data_start + i - 1
        if r != style_src:
            _copy_row_style(ws, style_src, r)
        ws.cell(r, 2).value = i
        ws.cell(r, 3).value = row["replacement"]
        ws.cell(r, 4).value = row["hours"]
        ws.cell(r, 5).value = row["dates"]
        ws.cell(r, 6).value = row["absent"]
        ws.cell(r, 7).value = row["reason"]
        _merge_gi(ws, r)

    sum_hours = sum(x["hours"] for x in groups)
    ws.cell(total_row, 4).value = sum_hours

    ws["A14"].value = _build_title(date_from, date_to)
    ws["C9"].value = datetime.combine(order_date, time(0, 0, 0))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
