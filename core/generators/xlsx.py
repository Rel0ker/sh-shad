"""Экспорт XLSX: учёт замен и таблицы для печати (учителя / ученики)."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="B9C4DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(bold=True, size=10)
_HEADER_FILL = PatternFill("solid", fgColor="EDF1F9")
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(size=11, color="45526B")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER


def _write_data_rows(
    ws,
    start_row: int,
    rows: list[dict[str, Any]],
    columns: list[tuple[str, str]],
    *,
    empty_text: str = "Нет данных на этот день",
) -> int:
    """Возвращает номер последней занятой строки."""
    if not rows:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=len(columns),
        )
        c = ws.cell(row=start_row, column=1, value=empty_text)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(italic=True, color="666666")
        for col in range(1, len(columns) + 1):
            ws.cell(row=start_row, column=col).border = _BORDER
        return start_row

    for r_idx, row in enumerate(rows, start=start_row):
        for col_idx, (key, _) in enumerate(columns, start=1):
            val = row.get(key, "")
            if val == "" or val is None:
                val = "—"
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.alignment = _LEFT if col_idx > 2 else _CENTER
            c.border = _BORDER
    return start_row + len(rows) - 1 if rows else start_row


def build_elementary_students_print_workbook(
    date_display: str,
    rows: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Нач. школа"

    _sheet_title_block(
        ws,
        "Изменения в расписании",
        date_display,
        "Начальная школа (1–4) · Для обучающихся",
    )

    headers = [
        ("klass", "Класс"),
        ("lesson_no", "№ урока"),
        ("subject", "Урок"),
        ("room", "Кабинет"),
        ("note", "Примечание"),
    ]
    hr = 5
    for col, (_, label) in enumerate(headers, start=1):
        ws.cell(row=hr, column=col, value=label)
    _style_header_row(ws, hr, len(headers))
    _write_data_rows(ws, hr + 1, rows, headers, empty_text="Нет изменений")

    widths = (12, 10, 36, 14, 24)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_title_block(ws, title: str, subtitle: str, badge: str) -> int:
    ws.merge_cells("A1:F1")
    ws["A1"] = badge
    ws["A1"].font = Font(size=9, bold=True, color="33415E")
    ws.merge_cells("A2:F2")
    ws["A2"] = title
    ws["A2"].font = _TITLE_FONT
    ws.merge_cells("A3:F3")
    ws["A3"] = subtitle
    ws["A3"].font = _SUBTITLE_FONT
    return 5


def build_teachers_print_workbook(
    date_display: str,
    rows: list[dict[str, Any]],
    *,
    badge: str = "Для учителей",
    sheet_title: str = "Учителя",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    _sheet_title_block(
        ws,
        "Изменения в расписании",
        date_display,
        badge,
    )

    headers = [
        ("klass", "Класс"),
        ("lesson_no", "№ урока"),
        ("absent_fio", "Отсутствующий учитель"),
        ("replacement_fio", "Заменяющий учитель"),
        ("room", "Кабинет"),
        ("note", "Примечание"),
    ]
    hr = 5
    for col, (_, label) in enumerate(headers, start=1):
        ws.cell(row=hr, column=col, value=label)
    _style_header_row(ws, hr, len(headers))

    last = _write_data_rows(ws, hr + 1, rows, headers)
    ws.freeze_panes = f"A{hr + 1}"
    if rows:
        ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"

    widths = (12, 10, 28, 28, 12, 24)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_students_print_workbook(
    date_display: str,
    rows_shift1: list[dict[str, Any]],
    rows_shift2: list[dict[str, Any]],
    *,
    badge: str = "Для обучающихся",
    sheet_title: str = "Ученики",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    _sheet_title_block(
        ws,
        "Изменения в расписании",
        date_display,
        badge,
    )

    headers = [
        ("klass", "Класс"),
        ("lesson_no", "№ урока"),
        ("subject", "Урок"),
        ("room", "Кабинет"),
        ("note", "Примечание"),
    ]
    col_count = len(headers)
    row = 5

    def write_shift_section(shift_title: str, shift_rows: list[dict[str, Any]]) -> int:
        nonlocal row
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=col_count,
        )
        c = ws.cell(row=row, column=1, value=shift_title)
        c.font = Font(bold=True, size=11, color="1F2A44")
        row += 1
        for col, (_, label) in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=label)
        _style_header_row(ws, row, col_count)
        row += 1
        empty = "Нет изменений для отображения"
        last = _write_data_rows(
            ws, row, shift_rows, headers, empty_text=empty
        )
        row = last + 2
        return row

    row = write_shift_section("1 смена", rows_shift1)
    write_shift_section("2 смена", rows_shift2)

    widths = (12, 10, 36, 14, 24)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_changes_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Замены"

    headers = [
        "Дата",
        "Отсутствующий",
        "Заменяющий",
        "Класс",
        "№ урока",
        "Предмет",
        "Кабинет",
        "Примечание",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, r in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=r.get("date"))
        ws.cell(row=r_idx, column=2, value=r.get("absent_fio"))
        ws.cell(row=r_idx, column=3, value=r.get("replacement_fio"))
        ws.cell(row=r_idx, column=4, value=r.get("klass"))
        ws.cell(row=r_idx, column=5, value=r.get("lesson_no"))
        ws.cell(row=r_idx, column=6, value=r.get("subject"))
        ws.cell(row=r_idx, column=7, value=r.get("room"))
        ws.cell(row=r_idx, column=8, value=r.get("note"))

    ws.freeze_panes = "A2"
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
