"""
Локальный веб-интерфейс для ввода изменений расписания и экспорта PDF/PNG/XLSX.
Запуск: python app.py  → откроется браузер на 127.0.0.1:<порт>
"""

from __future__ import annotations

import csv
import io
import os
import sys
import threading
import webbrowser
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)

from core import db
from core.export_data import (
    build_student_rows,
    build_teacher_rows,
    validate_lesson,
    xlsx_rows,
)
from core.generators import pdf as gen_pdf
from core.generators import png as gen_png
from core.generators import xlsx as gen_xlsx
from core.generators import xlsx_prikaz as gen_xlsx_prikaz
from core.util_dates import format_date_ru


def _resource_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent


def create_app() -> Flask:
    root = _resource_root()
    app = Flask(
        __name__,
        template_folder=str(root / "templates"),
        static_folder=str(root / "static"),
    )
    app.config["JSON_AS_ASCII"] = False

    @app.before_request
    def _ensure_db_once() -> None:
        if app.extensions.get("_schedule_db_inited"):
            return
        db.init_db()
        app.extensions["_schedule_db_inited"] = True

    def _css_url() -> str:
        return url_for("static", filename="app.css", _external=True)

    def _shift_for_payload_row(row: dict[str, Any]) -> int:
        cid = row.get("class_id")
        if cid is not None:
            s = db.class_shift_by_id(int(cid))
            if s:
                return s
        klass = (row.get("klass") or "").strip()
        if klass:
            s = db.class_shift_by_name(klass)
            if s:
                return s
        return 1

    def _validate_rows(rows: list[dict[str, Any]]) -> None:
        for i, r in enumerate(rows):
            shift = _shift_for_payload_row(r)
            try:
                ln = int(r["lesson_no"])
            except (KeyError, TypeError, ValueError):
                abort(400, description=f"Строка {i + 1}: неверный № урока")
            if not validate_lesson(shift, ln):
                lo, hi = (-1, 6) if shift == 2 else (0, 7)
                abort(
                    400,
                    description=(
                        f"Строка {i + 1}: для смены {shift} допустимы уроки {lo}…{hi}, "
                        f"указано {ln}"
                    ),
                )

    @app.get("/")
    def index() -> str:
        return render_template("index.html", default_date=db.today_iso())

    @app.get("/admin")
    def admin_page() -> str:
        return render_template("admin.html")

    @app.get("/api/reference")
    def api_reference() -> Any:
        return jsonify(
            {
                "teachers": db.list_teachers(),
                "classes": db.list_classes(),
                "subjects": db.list_subjects(),
            }
        )

    @app.post("/api/teachers")
    def api_add_teacher() -> Any:
        data = request.get_json(force=True, silent=True) or {}
        try:
            tid = db.add_teacher(str(data.get("fio", "")))
        except ValueError as e:
            abort(400, description=str(e))
        return jsonify({"id": tid}), 201

    @app.delete("/api/teachers/<int:tid>")
    def api_del_teacher(tid: int) -> Any:
        db.delete_teacher(tid)
        return "", 204

    @app.post("/api/classes")
    def api_add_class() -> Any:
        data = request.get_json(force=True, silent=True) or {}
        try:
            cid = db.add_class(str(data.get("name", "")), int(data.get("shift", 1)))
        except ValueError as e:
            abort(400, description=str(e))
        return jsonify({"id": cid}), 201

    @app.delete("/api/classes/<int:cid>")
    def api_del_class(cid: int) -> Any:
        db.delete_class(cid)
        return "", 204

    @app.post("/api/subjects")
    def api_add_subject() -> Any:
        data = request.get_json(force=True, silent=True) or {}
        try:
            sid = db.add_subject(str(data.get("name", "")))
        except ValueError as e:
            abort(400, description=str(e))
        return jsonify({"id": sid}), 201

    @app.delete("/api/subjects/<int:sid>")
    def api_del_subject(sid: int) -> Any:
        db.delete_subject(sid)
        return "", 204

    def _parse_spreadsheetml_rows(raw: bytes) -> list[list[str]]:
        """Парсер Excel 2003 XML (SpreadsheetML): вытаскивает строки листа."""
        root = ET.fromstring(raw)
        ns = {
            "ss": "urn:schemas-microsoft-com:office:spreadsheet",
        }
        rows: list[list[str]] = []
        for row in root.findall(".//ss:Worksheet/ss:Table/ss:Row", ns):
            cells: list[str] = []
            col = 1
            for cell in row.findall("ss:Cell", ns):
                idx_attr = cell.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
                if idx_attr and idx_attr.isdigit():
                    target = int(idx_attr)
                    while col < target:
                        cells.append("")
                        col += 1
                data = cell.find("ss:Data", ns)
                text = ""
                if data is not None and data.text is not None:
                    text = data.text.strip()
                cells.append(text)
                col += 1
            rows.append(cells)
        return rows

    def _import_first_column_fio(file_storage) -> int:
        from openpyxl import load_workbook

        name = (file_storage.filename or "").lower()
        raw = file_storage.read()
        n = 0
        if name.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace").splitlines()
            reader = csv.reader(text)
            for row in reader:
                if row and row[0].strip():
                    try:
                        db.add_teacher(row[0].strip())
                        n += 1
                    except ValueError:
                        pass
        elif name.endswith(".xlsx") or name.endswith(".xls"):
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and row[0] and str(row[0]).strip():
                    try:
                        db.add_teacher(str(row[0]).strip())
                        n += 1
                    except ValueError:
                        pass
            wb.close()
        elif name.endswith(".xml"):
            rows = _parse_spreadsheetml_rows(raw)
            for row in rows:
                if row and row[0].strip() and row[0].strip().lower() not in ("фио", "учитель"):
                    try:
                        db.add_teacher(row[0].strip())
                        n += 1
                    except ValueError:
                        pass
        else:
            abort(400, description="Ожидается CSV, XLSX или XML")
        return n

    def _import_classes_file(file_storage) -> int:
        from openpyxl import load_workbook

        name = (file_storage.filename or "").lower()
        raw = file_storage.read()
        n = 0
        if name.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace").splitlines()
            reader = csv.reader(text)
            for row in reader:
                if not row or not str(row[0]).strip():
                    continue
                sh = 1
                if len(row) > 1 and str(row[1]).strip().isdigit():
                    sh = int(str(row[1]).strip())
                    if sh not in (1, 2):
                        sh = 1
                try:
                    db.add_class(str(row[0]).strip(), sh)
                    n += 1
                except ValueError:
                    pass
        elif name.endswith(".xlsx") or name.endswith(".xls"):
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or not row[0] or not str(row[0]).strip():
                    continue
                sh = 1
                if len(row) > 1 and row[1] is not None and str(row[1]).strip().isdigit():
                    sh = int(str(row[1]).strip())
                    if sh not in (1, 2):
                        sh = 1
                try:
                    db.add_class(str(row[0]).strip(), sh)
                    n += 1
                except ValueError:
                    pass
            wb.close()
        elif name.endswith(".xml"):
            rows = _parse_spreadsheetml_rows(raw)
            for row in rows:
                if not row:
                    continue
                class_name = str(row[0]).strip()
                if not class_name or class_name.lower() == "класс":
                    continue
                sh = 2 if class_name.startswith(("1", "2", "3", "4")) else 1
                try:
                    db.add_class(class_name, sh)
                    n += 1
                except ValueError:
                    pass
        else:
            abort(400, description="Ожидается CSV, XLSX или XML")
        return n

    def _import_subjects_file(file_storage) -> int:
        from openpyxl import load_workbook

        name = (file_storage.filename or "").lower()
        raw = file_storage.read()
        subjects: set[str] = set()
        teachers: set[str] = set()
        if name.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace").splitlines()
            reader = csv.reader(text)
            for row in reader:
                if len(row) >= 3:
                    val = str(row[2]).strip()
                    if val and val.lower() != "предмет":
                        subjects.add(val)
                if len(row) >= 4:
                    tval = str(row[3]).strip()
                    if tval and tval.lower() != "учитель":
                        teachers.add(tval)
        elif name.endswith(".xlsx") or name.endswith(".xls"):
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and len(row) >= 3 and row[2]:
                    val = str(row[2]).strip()
                    if val and val.lower() != "предмет":
                        subjects.add(val)
                if row and len(row) >= 4 and row[3]:
                    tval = str(row[3]).strip()
                    if tval and tval.lower() != "учитель":
                        teachers.add(tval)
            wb.close()
        elif name.endswith(".xml"):
            rows = _parse_spreadsheetml_rows(raw)
            for row in rows:
                if len(row) >= 3:
                    val = str(row[2]).strip()
                    if val and val.lower() != "предмет":
                        subjects.add(val)
                if len(row) >= 4:
                    tval = str(row[3]).strip()
                    if tval and tval.lower() != "учитель":
                        teachers.add(tval)
        else:
            abort(400, description="Ожидается CSV, XLSX или XML")
        inserted_subjects = db.upsert_subjects(sorted(subjects))
        db.upsert_teachers(sorted(teachers))
        return inserted_subjects

    @app.post("/admin/import/teachers")
    def import_teachers() -> Any:
        f = request.files.get("file")
        if not f or not f.filename:
            return redirect(url_for("admin_page"))
        n = _import_first_column_fio(f)
        return redirect(url_for("admin_page") + f"?imported_teachers={n}")

    @app.post("/admin/import/classes")
    def import_classes() -> Any:
        f = request.files.get("file")
        if not f or not f.filename:
            return redirect(url_for("admin_page"))
        n = _import_classes_file(f)
        return redirect(url_for("admin_page") + f"?imported_classes={n}")

    @app.post("/admin/import/subjects")
    def import_subjects() -> Any:
        f = request.files.get("file")
        if not f or not f.filename:
            return redirect(url_for("admin_page"))
        n = _import_subjects_file(f)
        return redirect(url_for("admin_page") + f"?imported_subjects={n}")

    @app.get("/api/changes")
    def api_get_changes() -> Any:
        d = request.args.get("date") or db.today_iso()
        try:
            date.fromisoformat(d)
        except ValueError:
            abort(400, description="Неверная дата")
        rows = db.load_changes_for_day(d)
        return jsonify({"date": d, "rows": rows})

    @app.post("/api/changes")
    def api_post_changes() -> Any:
        data = request.get_json(force=True, silent=True) or {}
        d = str(data.get("date") or "")
        try:
            date.fromisoformat(d)
        except ValueError:
            abort(400, description="Неверная дата")
        rows = data.get("rows") or []
        if not isinstance(rows, list):
            abort(400, description="rows должен быть массивом")
        _validate_rows(rows)
        day_id = db.get_or_create_day(d)
        db.replace_changes_for_day(day_id, rows)
        return jsonify({"ok": True, "date": d, "count": len(rows)})

    def _changes_for_date(d: str) -> list[dict[str, Any]]:
        rows = db.load_changes_for_day(d)
        return rows

    @app.get("/print/teachers")
    def print_teachers() -> str:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        rows = build_teacher_rows(_changes_for_date(d))
        return render_template(
            "print_teachers.html",
            date_display=format_date_ru(d),
            rows=rows,
            css_url=_css_url(),
        )

    @app.get("/print/students")
    def print_students() -> str:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        s1, s2 = build_student_rows(_changes_for_date(d))
        return render_template(
            "print_students.html",
            date_display=format_date_ru(d),
            rows_shift1=s1,
            rows_shift2=s2,
            css_url=_css_url(),
        )

    def _export_error_response(exc: BaseException) -> Response:
        return Response(str(exc), status=503, mimetype="text/plain; charset=utf-8")

    @app.get("/export/xlsx/teachers")
    def export_xlsx_teachers() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        rows = build_teacher_rows(_changes_for_date(d))
        data = gen_xlsx.build_teachers_print_workbook(format_date_ru(d), rows)
        return send_file(
            io.BytesIO(data),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=f"замещение_учителей_{d}.xlsx",
        )

    @app.get("/export/xlsx/students")
    def export_xlsx_students() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        s1, s2 = build_student_rows(_changes_for_date(d))
        data = gen_xlsx.build_students_print_workbook(format_date_ru(d), s1, s2)
        return send_file(
            io.BytesIO(data),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=f"изменения_ученики_{d}.xlsx",
        )

    @app.get("/export/pdf/teachers")
    def export_pdf_teachers() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        rows = build_teacher_rows(_changes_for_date(d))
        ctx = {
            "date_display": format_date_ru(d),
            "rows": rows,
            "css_url": _css_url(),
        }
        base = request.host_url.rstrip("/")
        fb = f"{base}{url_for('print_teachers')}?date={d}"
        try:
            data = gen_pdf.render_pdf_bytes(
                app,
                "print_teachers.html",
                ctx,
                fallback_url=fb,
                fallback_landscape=True,
                fallback_margin_mm=6,
            )
        except Exception as e:
            return _export_error_response(e)
        return send_file(
            io.BytesIO(data),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"замещение_учителей_{d}.pdf",
        )

    @app.get("/export/pdf/students")
    def export_pdf_students() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        s1, s2 = build_student_rows(_changes_for_date(d))
        ctx = {
            "date_display": format_date_ru(d),
            "rows_shift1": s1,
            "rows_shift2": s2,
            "css_url": _css_url(),
        }
        base = request.host_url.rstrip("/")
        fb = f"{base}{url_for('print_students')}?date={d}"
        try:
            data = gen_pdf.render_pdf_bytes(
                app,
                "print_students.html",
                ctx,
                fallback_url=fb,
                fallback_landscape=False,
                fallback_margin_mm=8,
            )
        except Exception as e:
            return _export_error_response(e)
        return send_file(
            io.BytesIO(data),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"изменения_ученики_{d}.pdf",
        )

    @app.get("/export/png/teachers")
    def export_png_teachers() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        base = request.host_url.rstrip("/")
        url = f"{base}{url_for('print_teachers')}?date={d}"
        try:
            png = gen_png.screenshot_element(url)
        except Exception as e:
            return _export_error_response(e)
        return send_file(
            io.BytesIO(png),
            mimetype="image/png",
            as_attachment=True,
            download_name=f"замещение_учителей_{d}.png",
        )

    @app.get("/export/png/students")
    def export_png_students() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        base = request.host_url.rstrip("/")
        url = f"{base}{url_for('print_students')}?date={d}"
        try:
            png = gen_png.screenshot_element(url)
        except Exception as e:
            return _export_error_response(e)
        return send_file(
            io.BytesIO(png),
            mimetype="image/png",
            as_attachment=True,
            download_name=f"изменения_ученики_{d}.png",
        )

    @app.get("/export/xlsx/prikaz")
    def export_xlsx_prikaz() -> Any:
        d1 = request.args.get("date_from") or ""
        d2 = request.args.get("date_to") or ""
        try:
            df = date.fromisoformat(d1)
            dt = date.fromisoformat(d2)
        except ValueError:
            abort(400, description="Укажите date_from и date_to в формате YYYY-MM-DD")
        if df > dt:
            abort(400, description="date_from не может быть позже date_to")
        od = request.args.get("order_date")
        order_d: date | None = None
        if od:
            try:
                order_d = date.fromisoformat(str(od))
            except ValueError:
                abort(400, description="Неверный order_date")
        changes = db.load_changes_between(d1, d2)
        try:
            data = gen_xlsx_prikaz.build_prikaz_xlsx(
                changes, d1, d2, order_date=order_d
            )
        except FileNotFoundError as e:
            abort(500, description=str(e))
        fn = f"Замены_{d1}_{d2}.xlsx"
        return send_file(
            io.BytesIO(data),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=fn,
        )

    @app.get("/export/xlsx")
    def export_xlsx() -> Any:
        d = request.args.get("date") or db.today_iso()
        date.fromisoformat(d)
        rows_raw = _changes_for_date(d)
        rows = xlsx_rows(rows_raw, d)
        data = gen_xlsx.build_changes_workbook(rows)
        return send_file(
            io.BytesIO(data),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=f"учет_замен_{d}.xlsx",
        )

    return app


def _find_free_port(start: int = 8765, attempts: int = 40) -> int:
    import socket

    for p in range(start, start + attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("127.0.0.1", p))
            except OSError:
                continue
            return p
    return start


def main() -> None:
    os.environ.setdefault("FLASK_ENV", "development")
    app = create_app()
    port = int(os.environ.get("PORT", "0")) or _find_free_port()

    def _open_browser() -> None:
        webbrowser.open(f"http://127.0.0.1:{port}/")

    threading.Timer(0.8, _open_browser).start()
    # threaded=True — Playwright может дернуть /print/* пока обрабатывается /export/png/*
    app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    main()
